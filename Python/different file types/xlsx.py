from openpyxl import Workbook, load_workbook         # import excel main module
from openpyxl.styles import Font, Fill, PatternFill  # import excel styles


def save_xlsx():
    """
    Create xlsx file in the folder where the executable file is located
    Title default active datasheet colored into red and renamed
    A4 = 50
    B1:B19 = [12, 24, 36, 48, 60, 72, 84, 96, 108, 120, 132, 144, 156, 168, 180, 192, 204, 216, 228]
    """
    workbook = Workbook()                           # init
    worksheet = workbook.active                     # get default worksheet
    worksheet.sheet_properties.tabColor = "FF0000"  # paint over title
    worksheet.title = "NewTitle"                    # rename title
    worksheet['A4'] = 500                           # set single value to cell A4
    new_font = Font(size=23, underline='single', color='FFBB00', bold=True, italic=True)           # create text style
    worksheet['A4'].font = new_font                                                                # change text style
    new_cell_color = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')  # create cell color
    worksheet['A4'].fill = new_cell_color                                                          # change cell color
    for i in range(1, 20):                            # set field B1:B20 values from 12 to 228 = 12 * 19
        worksheet.cell(row=i, column=2, value=i * 12)
    workbook.save('1.xlsx')                           # save workbook


def load_xlsx():
    """
    Load xlsx file in the folder where the executable file is located
    Get single value from cell
    Get values from area
    """
    workbook = load_workbook('1.xlsx')  # load xlsx file and create workbook object
    sheet = workbook['NewTitle']        # get datasheet by title name
    print(sheet['A4'].value)            # get value A4 cell
    for cellObj in sheet['B1':'B10']:   # get location and values from area B1:B10
        for cell in cellObj:
            print(cell.coordinate, cell.value)


if __name__ == '__main__':
    save_xlsx()
    load_xlsx()
